"""Generates the .xlsx NOL tracker from the saved records.

Built fresh from the JSON store on every request so the spreadsheet is
always in sync with the data — reliable even on hosts with an ephemeral
filesystem.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import NOL_FIELDS, FIELD_LABELS

MTA_NAVY = "003DA5"
THIN = Side(style="thin", color="EAECF0")


def build_workbook(records) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "NOLs"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color=MTA_NAVY)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Header row
    for col, field in enumerate(NOL_FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=FIELD_LABELS[field])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for r, record in enumerate(records, start=2):
        for c, field in enumerate(NOL_FIELDS, start=1):
            cell = ws.cell(row=r, column=c, value=record.get(field, ""))
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    for col, field in enumerate(NOL_FIELDS, start=1):
        width = max(len(FIELD_LABELS[field]) + 3, 14)
        if field in ("Violation_Location", "Remarks"):
            width = 32
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
